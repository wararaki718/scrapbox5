from datetime import datetime

import pandas as pd
from openpyxl import Workbook


def main():
    wb = Workbook()
    filename = "sample.xlsx"

    ws = wb.active
    ws["A1"] = 12
    ws.append([1, 2, 3, 4])
    ws["A2"] = datetime.now()
    wb.save(filename)

    df = pd.read_excel(filename)
    print(df)
    
    df.to_excel(filename, index=None, header=None)
    print("DONE")


if __name__ == "__main__":
    main()
